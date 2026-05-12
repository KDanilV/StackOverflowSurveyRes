"""Базовый EDA и отбор важных полей для многолетних выгрузок Stack Overflow."""

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from stackoverflow_analytics.config import MULTIYEAR_PROCESSED_DIR, RAW_DATA_DIR

YEARS = (2022, 2023, 2024, 2025)

CORE_COLUMNS = [
    "ResponseId",
    "MainBranch",
    "Age",
    "Employment",
    "RemoteWork",
    "EdLevel",
    "YearsCode",
    "YearsCodePro",
    "WorkExp",
    "DevType",
    "OrgSize",
    "Country",
    "Currency",
    "CompTotal",
    "ConvertedCompYearly",
    "Industry",
    "AISelect",
    "AISent",
    "AIAcc",
    "AIThreat",
    "JobSat",
]

JOB_SAT_POINT_COLUMNS = [f"JobSatPoints_{index}" for index in range(1, 17)]

AI_ADOPTION_COLUMNS = [
    "AIToolCurrently Using",
    "AIToolCurrently mostly AI",
    "AIToolCurrently partially AI",
    "AIToolPlan to mostly use AI",
    "AIToolPlan to partially use AI",
    "AIToolInterested in Using",
    "AIToolNot interested in Using",
    "AIToolDon't plan to use AI for this task",
    "AIAgents",
    "AIAgentChange",
]

TECH_ENDORSEMENT_COLUMNS = [
    *[f"TechEndorse_{index}" for index in range(1, 10)],
    "TechEndorse_13",
    *[f"TechOppose_{index}" for index in range(1, 17)],
]

TECHNOLOGY_COLUMNS = {
    "language": (
        "LanguageHaveWorkedWith",
        "LanguageWantToWorkWith",
        "LanguageAdmired",
    ),
    "database": (
        "DatabaseHaveWorkedWith",
        "DatabaseWantToWorkWith",
        "DatabaseAdmired",
    ),
    "platform": (
        "PlatformHaveWorkedWith",
        "PlatformWantToWorkWith",
        "PlatformAdmired",
    ),
    "web_framework": (
        "WebframeHaveWorkedWith",
        "WebframeWantToWorkWith",
        "WebframeAdmired",
    ),
    "other_framework": (
        "MiscTechHaveWorkedWith",
        "MiscTechWantToWorkWith",
        "MiscTechAdmired",
    ),
    "developer_tool": (
        "ToolsTechHaveWorkedWith",
        "ToolsTechWantToWorkWith",
        "ToolsTechAdmired",
        "DevEnvsHaveWorkedWith",
        "DevEnvsWantToWorkWith",
        "DevEnvsAdmired",
    ),
    "collaboration_tool": (
        "NEWCollabToolsHaveWorkedWith",
        "NEWCollabToolsWantToWorkWith",
        "NEWCollabToolsAdmired",
        "CommPlatformHaveWorkedWith",
        "CommPlatformWantToWorkWith",
        "CommPlatformAdmired",
    ),
    "ai_search_or_dev": (
        "AISearchHaveWorkedWith",
        "AISearchWantToWorkWith",
        "AIDevHaveWorkedWith",
        "AIDevWantToWorkWith",
        "AISearchDevHaveWorkedWith",
        "AISearchDevWantToWorkWith",
        "AISearchDevAdmired",
        "AIModelsHaveWorkedWith",
        "AIModelsWantToWorkWith",
        "AIModelsAdmired",
    ),
    "stack_overflow_tag": (
        "SOTagsHaveWorkedWith",
        "SOTagsWantToWorkWith",
        "SOTagsAdmired",
    ),
}


@dataclass(frozen=True)
class YearSource:
    year: int
    path: Path


def source_for_year(year: int) -> YearSource:
    if year == 2024:
        return YearSource(year=year, path=RAW_DATA_DIR / "survey_results_public.xlsx")
    return YearSource(year=year, path=RAW_DATA_DIR / str(year) / "survey_results_public.csv")


def read_columns(source: YearSource, columns: list[str] | None = None) -> pd.DataFrame:
    if not source.path.exists():
        raise FileNotFoundError(f"Не найден файл данных за {source.year}: {source.path}")

    if source.path.suffix.lower() == ".xlsx":
        frame = read_xlsx_columns(source.path, columns)
    else:
        frame = pd.read_csv(source.path, usecols=columns, low_memory=False)
    return frame.replace(r"^\s*$", pd.NA, regex=True).replace("NA", pd.NA)


def read_header(source: YearSource) -> list[str]:
    if source.path.suffix.lower() == ".xlsx":
        workbook = load_workbook(source.path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            header = next(worksheet.iter_rows(values_only=True), ())
            return [str(column) for column in header if column is not None]
        finally:
            workbook.close()
    return list(pd.read_csv(source.path, nrows=0).columns)


def read_xlsx_columns(path: Path, columns: list[str] | None = None) -> pd.DataFrame:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows_data = []

    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        header = list(next(rows, ()))
        wanted_columns = columns or [str(column) for column in header if column is not None]
        indexes = {
            column_name: header.index(column_name)
            for column_name in wanted_columns
            if column_name in header
        }

        for row in rows:
            rows_data.append(
                {
                    column_name: row[index] if index < len(row) else None
                    for column_name, index in indexes.items()
                }
            )
    finally:
        workbook.close()

    return pd.DataFrame(rows_data, columns=list(indexes))


def available_columns(header: list[str], candidates: list[str]) -> list[str]:
    return [column for column in candidates if column in header]


def normalize_years(value: Any) -> float | None:
    if pd.isna(value):
        return None
    if isinstance(value, str):
        value = value.strip()
        if value.startswith("Less than"):
            return 0.5
        if value.startswith("More than"):
            return 51.0
    number = pd.to_numeric(value, errors="coerce")
    if pd.isna(number):
        return None
    return float(number)


def build_core_frame(year: int, frame: pd.DataFrame) -> pd.DataFrame:
    core = frame.reindex(columns=CORE_COLUMNS).copy()
    core.insert(0, "SurveyYear", year)

    for column in ("YearsCode", "YearsCodePro", "WorkExp", "CompTotal", "ConvertedCompYearly"):
        if column in core.columns:
            core[column] = core[column].apply(normalize_years)

    core["HasSalary"] = core["ConvertedCompYearly"].notna()
    professional_experience = core["YearsCodePro"].copy()
    core["ProfessionalExperience"] = professional_experience.where(
        professional_experience.notna(), core["WorkExp"]
    )
    core["JobSatNormalized"] = normalize_job_satisfaction(frame)
    core["AIAdoption"] = normalize_ai_adoption(frame)
    return core


def normalize_job_satisfaction(frame: pd.DataFrame) -> pd.Series:
    if "JobSat" in frame:
        job_sat = pd.to_numeric(frame["JobSat"], errors="coerce")
        if job_sat.notna().any():
            return job_sat

    point_columns = [column for column in JOB_SAT_POINT_COLUMNS if column in frame]
    if not point_columns:
        return pd.Series(pd.NA, index=frame.index, dtype="Float64")

    points = frame[point_columns].apply(pd.to_numeric, errors="coerce")
    return points.mean(axis=1)


def normalize_ai_adoption(frame: pd.DataFrame) -> pd.Series:
    result = pd.Series("Нет данных", index=frame.index, dtype="object")

    if "AISelect" in frame:
        ai_select = frame["AISelect"].fillna("").astype(str).str.lower()
        result = result.mask(ai_select.str.contains("yes", regex=False), "Использует AI")
        result = result.mask(
            ai_select.str.contains("plan", regex=False), "Планирует использовать AI"
        )
        result = result.mask(ai_select.str.contains("no", regex=False), "Не использует AI")

    current_columns = [
        "AIToolCurrently Using",
        "AIToolCurrently mostly AI",
        "AIToolCurrently partially AI",
    ]
    planned_columns = [
        "AIToolPlan to mostly use AI",
        "AIToolPlan to partially use AI",
        "AIToolInterested in Using",
    ]
    rejected_columns = [
        "AIToolNot interested in Using",
        "AIToolDon't plan to use AI for this task",
    ]

    result = result.mask(has_any_value(frame, rejected_columns), "Не использует AI")
    result = result.mask(has_any_value(frame, planned_columns), "Планирует использовать AI")
    result = result.mask(has_any_value(frame, current_columns), "Использует AI")
    return result


def has_any_value(frame: pd.DataFrame, columns: list[str]) -> pd.Series:
    existing_columns = [column for column in columns if column in frame]
    if not existing_columns:
        return pd.Series(False, index=frame.index)
    return frame[existing_columns].notna().any(axis=1)


def split_multi_value(value: Any) -> list[str]:
    if pd.isna(value):
        return []
    return [item.strip() for item in str(value).split(";") if item.strip()]


def build_technology_frame(year: int, frame: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []

    for category, columns in TECHNOLOGY_COLUMNS.items():
        for column in columns:
            if column not in frame.columns:
                continue

            intent = technology_intent(column)
            for response_id, value in frame[["ResponseId", column]].itertuples(index=False):
                for technology in split_multi_value(value):
                    rows.append(
                        {
                            "SurveyYear": year,
                            "ResponseId": response_id,
                            "Category": category,
                            "Intent": intent,
                            "Technology": technology,
                            "SourceColumn": column,
                        }
                    )

    return pd.DataFrame(
        rows,
        columns=["SurveyYear", "ResponseId", "Category", "Intent", "Technology", "SourceColumn"],
    )


def technology_intent(column: str) -> str:
    if "Want" in column:
        return "wanted"
    if "Admired" in column:
        return "admired"
    return "worked"


def aggregate_technology_counts(technology_frames: list[pd.DataFrame]) -> pd.DataFrame:
    non_empty_frames = [frame for frame in technology_frames if not frame.empty]
    if not non_empty_frames:
        return pd.DataFrame(
            columns=["SurveyYear", "Category", "Intent", "Technology", "Respondents"]
        )

    technologies = pd.concat(non_empty_frames, ignore_index=True)
    if technologies.empty:
        return pd.DataFrame(
            columns=["SurveyYear", "Category", "Intent", "Technology", "Respondents"]
        )

    return (
        technologies.groupby(["SurveyYear", "Category", "Intent", "Technology"], as_index=False)
        .agg(Respondents=("ResponseId", "nunique"))
        .sort_values(
            ["SurveyYear", "Category", "Intent", "Respondents"], ascending=[True, True, True, False]
        )
    )


def summarize_year(year: int, header: list[str], frame: pd.DataFrame) -> dict[str, Any]:
    return {
        "year": year,
        "rows": int(len(frame)),
        "columns_total": int(len(header)),
        "columns_kept_core": int(len(available_columns(header, CORE_COLUMNS))),
        "countries": int(frame["Country"].nunique(dropna=True)) if "Country" in frame else 0,
        "salary_non_null": int(frame["ConvertedCompYearly"].notna().sum())
        if "ConvertedCompYearly" in frame
        else 0,
        "ai_columns": sorted(column for column in header if column.startswith("AI")),
        "kept_columns": available_columns(header, CORE_COLUMNS),
    }


def write_outputs(
    core_frames: list[pd.DataFrame],
    technology_frames: list[pd.DataFrame],
    summaries: list[dict[str, Any]],
    output_dir: Path,
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)

    core = pd.concat(
        [frame.dropna(axis=1, how="all") for frame in core_frames],
        ignore_index=True,
    ).reindex(
        columns=[
            "SurveyYear",
            *CORE_COLUMNS,
            "HasSalary",
            "ProfessionalExperience",
            "JobSatNormalized",
            "AIAdoption",
        ]
    )
    technology_counts = aggregate_technology_counts(technology_frames)

    core.to_csv(output_dir / "core_survey.csv", index=False)
    technology_counts.to_csv(output_dir / "technology_counts.csv", index=False)
    (output_dir / "eda_summary.json").write_text(
        json.dumps(summaries, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def run_eda(years: tuple[int, ...] = YEARS, output_dir: Path = MULTIYEAR_PROCESSED_DIR) -> None:
    core_frames = []
    technology_frames = []
    summaries = []

    for year in years:
        source = source_for_year(year)
        header = read_header(source)
        selected_columns = available_columns(
            header,
            sorted(
                set(CORE_COLUMNS)
                | set(JOB_SAT_POINT_COLUMNS)
                | set(AI_ADOPTION_COLUMNS)
                | set(TECH_ENDORSEMENT_COLUMNS)
                | {column for values in TECHNOLOGY_COLUMNS.values() for column in values}
            ),
        )
        frame = read_columns(source, selected_columns)
        summaries.append(summarize_year(year, header, frame))
        core_frames.append(build_core_frame(year, frame))
        technology_frames.append(build_technology_frame(year, frame))

    write_outputs(core_frames, technology_frames, summaries, output_dir)
    print(f"Готово. Multi-year EDA сохранен в {output_dir}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--years",
        nargs="+",
        type=int,
        default=list(YEARS),
        choices=list(YEARS),
        help="Годы для анализа.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=MULTIYEAR_PROCESSED_DIR,
        help="Папка для сгенерированных EDA-файлов.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    run_eda(tuple(args.years), args.output_dir)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
