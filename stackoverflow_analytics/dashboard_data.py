import pandas as pd
from openpyxl import load_workbook

from stackoverflow_analytics.config import (
    CLEANED_SURVEY_FILE,
    DASHBOARD_CORE_FILE,
    DASHBOARD_TECHNOLOGY_COUNTS_FILE,
    MULTIYEAR_CORE_FILE,
    MULTIYEAR_TECHNOLOGY_COUNTS_FILE,
)
from stackoverflow_analytics.tousd import calculate_salary_usd

MAIN_COLUMNS = [
    "ResponseId",
    "Country",
    "Age",
    "EdLevel",
    "RemoteWork",
    "DevType",
    "YearsCode",
    "YearsCodePro",
    "Currency",
    "CompTotal",
    "ConvertedCompYearly",
    "AISelect",
    "AISent",
    "JobSat",
]

TECHNOLOGY_SHEETS = {
    "Databases": "DatabaseHaveWorkedWith",
    "Platforms": "PlatformHaveWorkedWith",
    "Web frameworks": "WebframeHaveWorkedWith",
    "Developer tools": "ToolsTechHaveWorkedWith",
}


def read_sheet_columns(input_file, sheet_name, columns):
    if not input_file.exists():
        raise FileNotFoundError(f"Cleaned data file not found: {input_file}")

    workbook = load_workbook(input_file, read_only=True, data_only=True)
    rows_data = []

    try:
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return pd.DataFrame(columns=columns)

        indexes = {
            column_name: header.index(column_name)
            for column_name in columns
            if column_name in header
        }

        for row in rows:
            rows_data.append(
                {
                    column_name: row[index] if index < len(row) else None
                    for column_name, index in indexes.items()
                }
            )
    finally:
        workbook.close()

    return pd.DataFrame(rows_data, columns=list(indexes))


def read_main_table(input_file=CLEANED_SURVEY_FILE):
    return read_sheet_columns(input_file, "Main", MAIN_COLUMNS)


def prepare_main_table(main):
    main = main.copy()
    main["YearsCode"] = pd.to_numeric(main.get("YearsCode"), errors="coerce")
    main["YearsCodePro"] = pd.to_numeric(main.get("YearsCodePro"), errors="coerce")
    main["ConvertedCompYearly"] = pd.to_numeric(main.get("ConvertedCompYearly"), errors="coerce")

    salary = calculate_salary_usd(main[["ResponseId", "Currency", "CompTotal"]])
    salary = salary[["ResponseId", "salaryusd"]]

    return main.merge(salary, on="ResponseId", how="left")


def read_language_table(input_file=CLEANED_SURVEY_FILE):
    return read_sheet_columns(
        input_file,
        "LanguageHaveWorkedWith",
        ["ResponseId", "LanguageHaveWorkedWith"],
    )


def read_wanted_language_table(input_file=CLEANED_SURVEY_FILE):
    return read_sheet_columns(
        input_file,
        "LanguageWantToWorkWith",
        ["ResponseId", "LanguageWantToWorkWith"],
    )


def read_technology_tables(input_file=CLEANED_SURVEY_FILE):
    return {
        label: read_sheet_columns(input_file, sheet_name, ["ResponseId", sheet_name])
        for label, sheet_name in TECHNOLOGY_SHEETS.items()
    }


def filter_main_table(main, countries=None, remote_work=None, education=None):
    filtered = main.copy()

    if countries:
        filtered = filtered[filtered["Country"].isin(countries)]
    if remote_work:
        filtered = filtered[filtered["RemoteWork"].isin(remote_work)]
    if education:
        filtered = filtered[filtered["EdLevel"].isin(education)]

    return filtered


def split_multiselect_values(values: pd.Series) -> pd.Series:
    split_values = values.fillna("").astype(str).str.split(";").explode().str.strip()
    return split_values[split_values != ""]


def contains_any_multiselect_value(values: pd.Series, selected_values: list[str]) -> pd.Series:
    selected = set(selected_values)
    split_values = values.fillna("").astype(str).str.split(";")
    return split_values.apply(lambda items: bool({item.strip() for item in items} & selected))


def filter_multiyear_core(
    core,
    years=None,
    countries=None,
    remote_work=None,
    education=None,
    developer_type=None,
):
    filtered = core.copy()

    if filtered.empty:
        return filtered

    if years is not None:
        if not years:
            return filtered.iloc[0:0].copy()
        if "SurveyYear" in filtered:
            filtered = filtered[filtered["SurveyYear"].isin(years)]

    if countries and "Country" in filtered:
        filtered = filtered[filtered["Country"].isin(countries)]
    if remote_work and "RemoteWork" in filtered:
        filtered = filtered[filtered["RemoteWork"].isin(remote_work)]
    if education and "EdLevel" in filtered:
        filtered = filtered[filtered["EdLevel"].isin(education)]
    if developer_type and "DevType" in filtered:
        filtered = filtered[contains_any_multiselect_value(filtered["DevType"], developer_type)]

    return filtered


def filter_by_response_ids(frame, response_ids):
    if response_ids is None or "ResponseId" not in frame.columns:
        return frame
    return frame[frame["ResponseId"].isin(response_ids)]


def top_counts(frame, column, limit=15):
    if column not in frame.columns:
        return pd.DataFrame(columns=[column, "count"])

    return (
        frame[column]
        .dropna()
        .value_counts()
        .head(limit)
        .rename_axis(column)
        .reset_index(name="count")
    )


def median_salary_by_group(main, group_column, limit=20):
    if group_column not in main.columns:
        return pd.DataFrame(columns=[group_column, "median_salary_usd", "respondents"])

    salary = main.dropna(subset=[group_column, "salaryusd"])
    if salary.empty:
        return pd.DataFrame(columns=[group_column, "median_salary_usd", "respondents"])

    return (
        salary.groupby(group_column, as_index=False)
        .agg(median_salary_usd=("salaryusd", "median"), respondents=("ResponseId", "count"))
        .sort_values(["median_salary_usd", "respondents"], ascending=[False, False])
        .head(limit)
    )


def salary_by_country(main, limit=20):
    return median_salary_by_group(main, "Country", limit)


def median_compensation_by_group(core, group_column, limit=20):
    if group_column not in core.columns:
        return pd.DataFrame(columns=[group_column, "median_salary_usd", "respondents"])

    salary = core.dropna(subset=[group_column, "ConvertedCompYearly"])
    if salary.empty:
        return pd.DataFrame(columns=[group_column, "median_salary_usd", "respondents"])

    return (
        salary.groupby(group_column, as_index=False)
        .agg(
            median_salary_usd=("ConvertedCompYearly", "median"),
            respondents=("ResponseId", "count"),
        )
        .sort_values(["median_salary_usd", "respondents"], ascending=[False, False])
        .head(limit)
    )


def language_popularity(language, response_ids=None, limit=20):
    language = filter_by_response_ids(language, response_ids)

    value_column = "LanguageHaveWorkedWith"
    if value_column not in language.columns:
        return pd.DataFrame(columns=[value_column, "count"])

    return (
        language[value_column]
        .dropna()
        .value_counts()
        .head(limit)
        .rename_axis(value_column)
        .reset_index(name="count")
    )


def worked_vs_wanted_languages(worked, wanted, response_ids=None, limit=25):
    worked = filter_by_response_ids(worked, response_ids)
    wanted = filter_by_response_ids(wanted, response_ids)

    worked_counts = worked["LanguageHaveWorkedWith"].dropna().value_counts().rename("worked_count")
    wanted_counts = wanted["LanguageWantToWorkWith"].dropna().value_counts().rename("wanted_count")

    comparison = (
        pd.concat([worked_counts, wanted_counts], axis=1)
        .fillna(0)
        .astype(int)
        .rename_axis("language")
        .reset_index()
    )
    comparison["want_gap"] = comparison["wanted_count"] - comparison["worked_count"]

    return comparison.sort_values("wanted_count", ascending=False).head(limit)


def technology_popularity(technology_tables, response_ids=None, limit=15):
    frames = []

    for label, table in technology_tables.items():
        filtered = filter_by_response_ids(table, response_ids)
        value_columns = [column for column in filtered.columns if column != "ResponseId"]
        if not value_columns:
            continue

        value_column = value_columns[0]
        counts = top_counts(filtered, value_column, limit)
        counts = counts.rename(columns={value_column: "technology"})
        counts["category"] = label
        frames.append(counts)

    if not frames:
        return pd.DataFrame(columns=["technology", "count", "category"])

    return pd.concat(frames, ignore_index=True)


def technology_count_distribution(technology_tables, response_ids=None):
    counts = []

    for table in technology_tables.values():
        filtered = filter_by_response_ids(table, response_ids)
        counts.append(filtered[["ResponseId"]])

    if not counts:
        return pd.DataFrame(columns=["technology_count", "respondents"])

    all_technologies = pd.concat(counts, ignore_index=True)
    return (
        all_technologies["ResponseId"]
        .value_counts()
        .rename_axis("ResponseId")
        .reset_index(name="technology_count")["technology_count"]
        .value_counts()
        .sort_index()
        .rename_axis("technology_count")
        .reset_index(name="respondents")
    )


def read_multiyear_core(input_file=None):
    input_file = input_file or (
        DASHBOARD_CORE_FILE if DASHBOARD_CORE_FILE.exists() else MULTIYEAR_CORE_FILE
    )
    if not input_file.exists():
        return pd.DataFrame()

    core = pd.read_csv(input_file, low_memory=False)
    numeric_columns = [
        "SurveyYear",
        "ConvertedCompYearly",
        "ProfessionalExperience",
        "JobSatNormalized",
    ]
    for column in numeric_columns:
        if column in core:
            core[column] = pd.to_numeric(core[column], errors="coerce")
    return core


def read_multiyear_technology_counts(input_file=None):
    input_file = input_file or (
        DASHBOARD_TECHNOLOGY_COUNTS_FILE
        if DASHBOARD_TECHNOLOGY_COUNTS_FILE.exists()
        else MULTIYEAR_TECHNOLOGY_COUNTS_FILE
    )
    if not input_file.exists():
        return pd.DataFrame()

    counts = pd.read_csv(input_file)
    if "SurveyYear" in counts:
        counts["SurveyYear"] = pd.to_numeric(counts["SurveyYear"], errors="coerce")
    if "Respondents" in counts:
        counts["Respondents"] = pd.to_numeric(counts["Respondents"], errors="coerce")
    return counts


def filter_multiyear_years(frame, years=None):
    if frame.empty or not years or "SurveyYear" not in frame:
        return frame
    return frame[frame["SurveyYear"].isin(years)]


def multiyear_salary_trend(core, years=None):
    core = filter_multiyear_years(core, years)
    if core.empty:
        return pd.DataFrame(columns=["SurveyYear", "median_salary_usd", "respondents"])

    salary = core.dropna(subset=["SurveyYear", "ConvertedCompYearly"])
    return (
        salary.groupby("SurveyYear", as_index=False)
        .agg(
            median_salary_usd=("ConvertedCompYearly", "median"),
            respondents=("ResponseId", "count"),
        )
        .sort_values("SurveyYear")
    )


def multiyear_remote_trend(core, years=None):
    core = filter_multiyear_years(core, years)
    if core.empty or "RemoteWork" not in core:
        return pd.DataFrame(columns=["SurveyYear", "RemoteWork", "respondents"])

    return (
        core.dropna(subset=["SurveyYear", "RemoteWork"])
        .groupby(["SurveyYear", "RemoteWork"], as_index=False)
        .agg(respondents=("ResponseId", "count"))
        .sort_values(["SurveyYear", "respondents"], ascending=[True, False])
    )


def multiyear_ai_adoption_trend(core, years=None):
    core = filter_multiyear_years(core, years)
    if core.empty or "AIAdoption" not in core:
        return pd.DataFrame(columns=["SurveyYear", "AIAdoption", "respondents"])

    return (
        core.dropna(subset=["SurveyYear", "AIAdoption"])
        .groupby(["SurveyYear", "AIAdoption"], as_index=False)
        .agg(respondents=("ResponseId", "count"))
        .sort_values(["SurveyYear", "respondents"], ascending=[True, False])
    )


def multiyear_top_technologies(technology_counts, category, intent="worked", limit=10, years=None):
    technology_counts = filter_multiyear_years(technology_counts, years)
    if technology_counts.empty:
        return pd.DataFrame(
            columns=["SurveyYear", "Category", "Intent", "Technology", "Respondents"]
        )

    filtered = technology_counts[
        (technology_counts["Category"] == category) & (technology_counts["Intent"] == intent)
    ].copy()
    if filtered.empty:
        return filtered

    top_technologies = (
        filtered.groupby("Technology", as_index=False)["Respondents"]
        .sum()
        .sort_values("Respondents", ascending=False)
        .head(limit)["Technology"]
    )
    return filtered[filtered["Technology"].isin(top_technologies)].sort_values(
        ["SurveyYear", "Respondents"], ascending=[True, False]
    )


def salary_map_by_country(core, years=None, min_respondents=30):
    columns = ["Country", "median_salary_usd", "respondents"]
    required_columns = {"Country", "ConvertedCompYearly", "ResponseId"}
    if core.empty or not required_columns.issubset(core.columns):
        return pd.DataFrame(columns=columns)

    salary = core.copy()
    salary = filter_multiyear_years(salary, years)

    salary = salary.dropna(subset=["Country", "ConvertedCompYearly"])
    salary["Country"] = salary["Country"].astype(str).str.strip()
    salary = salary[salary["Country"] != ""]
    salary["ConvertedCompYearly"] = pd.to_numeric(salary["ConvertedCompYearly"], errors="coerce")
    salary = salary.dropna(subset=["ConvertedCompYearly"])
    if salary.empty:
        return pd.DataFrame(columns=columns)

    result = (
        salary.groupby("Country", as_index=False)
        .agg(
            median_salary_usd=("ConvertedCompYearly", "median"),
            respondents=("ResponseId", "count"),
        )
        .query("respondents >= @min_respondents")
        .sort_values("median_salary_usd", ascending=False)
    )
    return result
