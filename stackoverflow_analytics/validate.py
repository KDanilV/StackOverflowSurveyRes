import json

import pandas as pd
from openpyxl import load_workbook

from stackoverflow_analytics.config import CLEANED_SURVEY_FILE, CLEANED_SURVEY_MANIFEST_FILE


def validate_cleaned_tables(tables):
    errors = []

    main = tables.get("Main")
    if main is None:
        return ['Missing required sheet: "Main"']

    if "ResponseId" not in main.columns:
        return ['Missing required column in "Main": ResponseId']

    if main["ResponseId"].isna().any():
        errors.append('"Main" contains rows with empty ResponseId')

    duplicated_ids = main.loc[main["ResponseId"].duplicated(), "ResponseId"].dropna().unique()
    if len(duplicated_ids) > 0:
        preview = ", ".join(str(value) for value in duplicated_ids[:5])
        errors.append(f'"Main" contains duplicate ResponseId values: {preview}')

    main_ids = set(main["ResponseId"].dropna())

    for sheet_name, table in tables.items():
        if sheet_name == "Main" or "ResponseId" not in table.columns:
            continue

        if table["ResponseId"].isna().any():
            errors.append(f'"{sheet_name}" contains rows with empty ResponseId')

        child_ids = set(table["ResponseId"].dropna())
        missing_ids = sorted(child_ids - main_ids)
        if missing_ids:
            preview = ", ".join(str(value) for value in missing_ids[:5])
            errors.append(f'"{sheet_name}" references unknown ResponseId values: {preview}')

    return errors


def read_sheet_manifest(manifest_file=CLEANED_SURVEY_MANIFEST_FILE):
    if not manifest_file.exists():
        return {}

    manifest = json.loads(manifest_file.read_text(encoding="utf-8"))
    return manifest.get("sheets", {})


def read_cleaned_workbook(
    input_file=CLEANED_SURVEY_FILE, manifest_file=CLEANED_SURVEY_MANIFEST_FILE
):
    if not input_file.exists():
        raise FileNotFoundError(f"Cleaned data file not found: {input_file}")

    sheet_manifest = read_sheet_manifest(manifest_file)
    workbook = load_workbook(input_file, read_only=True, data_only=True)
    tables = {}

    try:
        for worksheet in workbook.worksheets:
            logical_name = sheet_manifest.get(worksheet.title, worksheet.title)
            rows = worksheet.iter_rows(values_only=True)
            header = next(rows, None)

            if not header or "ResponseId" not in header:
                tables[logical_name] = pd.DataFrame()
                continue

            response_id_index = header.index("ResponseId")
            response_ids = [row[response_id_index] for row in rows if response_id_index < len(row)]
            tables[logical_name] = pd.DataFrame({"ResponseId": response_ids})
    finally:
        workbook.close()

    return tables


def main():
    tables = read_cleaned_workbook()
    errors = validate_cleaned_tables(tables)

    if errors:
        joined = "\n- ".join(errors)
        raise ValueError(f"Cleaned data validation failed:\n- {joined}")

    print(f"Done. Cleaned data validation passed for {len(tables)} sheets.")


if __name__ == "__main__":
    main()
